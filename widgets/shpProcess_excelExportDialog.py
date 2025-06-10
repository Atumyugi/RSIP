import sys
import os
import os.path as osp

from ui.shp_process_dialog.excelExportDialog import Ui_excelExportDialog

from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtWidgets import QApplication, QDialog, QHBoxLayout, QVBoxLayout
from PyQt5.QtGui import QFont,QIcon

from qgis.core import QgsProject,QgsMapLayer,QgsMapLayerType,QgsVectorLayer,QgsWkbTypes,QgsLayerTreeGroup,QgsRasterLayer,QgsRectangle,QgsMapSettings,QgsCoordinateReferenceSystem
from qgis.gui import QgsMapCanvas,QgsProjectionSelectionDialog

from qfluentwidgets import (CardWidget, setTheme, Theme, IconWidget, BodyLabel,SubtitleLabel, CaptionLabel, PushButton,
                            TransparentToolButton, RoundMenu, Action,MessageBox)
from qfluentwidgets import FluentIcon as FIF

from yoyiUtils.qtTriggeredCommon import qtTriggeredCommonDialog
from yoyiUtils.yoyiFile import checkTifList
from yoyiUtils.yoyiTranslate import yoyiTrans
import yoyirs_rc

from openpyxl import Workbook,load_workbook

PROJECT = QgsProject.instance()

class ExcelExportDialogClass(Ui_excelExportDialog,QDialog):
    def __init__(self,yoyiTrs:yoyiTrans,parent=None):
        super().__init__(parent)
        self.parentWindow = parent
        self.parentMapCanvas : QgsMapCanvas = self.parentWindow.mapCanvas
        self.yoyiTrs = yoyiTrs
        self.setupUi(self)
        self.initMember()
        self.initUI()
        self.connectFunc()
    
    def initMember(self):
        self.shpPath = None
        self.pointExcelPath = None
        self.riskField = None
        self.changeType = None
        self.highLevelRiskDistance = None
        self.mediumLevelRiskDistance = None
        self.lowLevelRiskDistance = None
        self.resPath = None
    
    def initUI(self):
        self.setWindowFlags(self.windowFlags() &~ Qt.WindowContextHelpButtonHint)
        for layer in PROJECT.mapLayers().values():
            layer : QgsMapLayer
            if layer.type() == QgsMapLayerType.VectorLayer and layer.geometryType() == QgsWkbTypes.PolygonGeometry:
                self.selectShpCb.addItem(layer.name())
                self.selectDangduanCb.addItem(layer.name())
            elif layer.type() == QgsMapLayerType.RasterLayer:
                self.selectTifCb.addItem(layer.name())

        self.selectSaveFilePB.setIcon(FIF.MORE)
        self.selectExcelFilePB.setIcon(FIF.MORE)

        self.selectTifRb.setChecked(True)
        self.selectDangduanRb.setChecked(True)

        self.refreshFieldView()

    def connectFunc(self):

        self.selectShpCb.currentIndexChanged.connect(self.refreshFieldView)

        self.selectExcelFilePB.clicked.connect(lambda: qtTriggeredCommonDialog().addFileTriggered(self.excelLE,'excel',parent=self))
        self.selectSaveFilePB.clicked.connect(lambda: qtTriggeredCommonDialog().selectSaveFileTriggered(self.resLE,'excel',parent=self))
        self.runPB.clicked.connect(self.runPBClicked)
        self.cancelPB.clicked.connect(self.close)

    def refreshFieldView(self):
        if self.selectShpCb.count() > 0:
            currentLayerName = self.selectShpCb.currentText()
            currentLayer : QgsVectorLayer = PROJECT.mapLayersByName(currentLayerName)[0]

            self.riskNameCb.clear()
            self.changeTypeCb.clear()
            typeNameIndex = 0
            changeTypeIndex = 0
            for index,field in enumerate(currentLayer.fields()):
                self.riskNameCb.addItem(field.name())
                self.changeTypeCb.addItem(field.name())
                if field.name() == "typename":
                    typeNameIndex = index
                elif field.name() == "BHLX":
                    changeTypeIndex = index
            self.riskNameCb.setCurrentIndex(typeNameIndex)
            self.changeTypeCb.setCurrentIndex(changeTypeIndex)


    def runPBClicked(self):
        if self.excelLE.text() == "" or self.resLE.text() == "":
            MessageBox(self.yoyiTrs._translate('警告'), self.yoyiTrs._translate('地址非法'), self).exec_()
            return
        
        if self.selectShpCb.count() == 0:
            MessageBox(self.yoyiTrs._translate('警告'), self.yoyiTrs._translate('没有有效矢量'), self).exec_()
            return
        
        if self.selectTifRb.isChecked() and self.selectTifCb.count() == 0:
            MessageBox(self.yoyiTrs._translate('警告'), self.yoyiTrs._translate('没有有效影像'), self).exec_()
            return
        
        if self.selectDangduanRb.isChecked() and self.selectDangduanCb.count() == 0:
            MessageBox(self.yoyiTrs._translate('警告'), self.yoyiTrs._translate('没有有效矢量'), self).exec_()
            return

        currentLayerName = self.selectShpCb.currentText()
        currentLayerSource : QgsVectorLayer = PROJECT.mapLayersByName(currentLayerName)[0].source()

        if self.selectTifRb.isChecked():
            self.extraTifLayer = PROJECT.mapLayersByName(self.selectTifCb.currentText())[0]
        else:
            self.extraTifLayer = None
        
        if self.selectDangduanRb.isChecked():
            self.extraDangduanPath = PROJECT.mapLayersByName(self.selectDangduanCb.currentText())[0].source()
        else:
            self.extraDangduanPath = None
        
        pointExcelPath = self.excelLE.text()

        pointWb = load_workbook(pointExcelPath,read_only=True)
        pointWs = pointWb.active
        if pointWs.max_row < 3:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中点位过少！请至少包含两个点位", self).exec_()
            return
        
        first_row = list(pointWs.iter_rows(values_only=True))[0]
        kvIndex = -9999 #电压等级的列索引
        lineNameIndex = -9999 #线路名称的列索引
        pointIdIndex = -9999 #杆塔编号的列索引
        groupIndex = -9999 #所属班组的列索引
        xIndex = -9999 #经度列索引
        yIndex = -9999 #纬度列索引
        for colIndex,colItem in enumerate(first_row):
            if colItem == "电压等级":
                kvIndex = colIndex
            elif colItem == "线路名称":
                lineNameIndex = colIndex
            elif colItem == "杆塔编号":
                pointIdIndex = colIndex
            elif colItem == "所属班组":
                groupIndex = colIndex
            elif colItem == "经度":
                xIndex = colIndex
            elif colItem == "纬度":
                yIndex = colIndex
        if kvIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<电压等级>列", self).exec_()
            return
        if lineNameIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<线路名称>列", self).exec_()
            return
        if pointIdIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<杆塔编号>列", self).exec_()
            return
        if groupIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<所属班组>列", self).exec_()
            return
        if xIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<经度>列", self).exec_()
            return
        if yIndex < 0:
            MessageBox(self.yoyiTrs._translate('警告'), "excel中缺少<纬度>列", self).exec_()
            return
        
        highLevelRiskDistance = self.highLevelRiskDistanceSb.value()
        mediumLevelRiskDistance = self.mediumLevelRiskDistanceSb.value()
        lowLevelRiskDistance = self.lowLevelRiskDistanceSb.value()

        if not highLevelRiskDistance < mediumLevelRiskDistance < lowLevelRiskDistance:
            MessageBox(self.yoyiTrs._translate('警告'), "当前未满足高风险距离<中风险距离<低风险距离，请调整！", self).exec_()
            return
        
        
        self.shpPath = currentLayerSource
        self.pointExcelPath = self.excelLE.text()
        self.riskField = self.riskNameCb.currentText()
        self.changeType = self.changeTypeCb.currentText()
        self.highLevelRiskDistance = highLevelRiskDistance
        self.mediumLevelRiskDistance = mediumLevelRiskDistance
        self.lowLevelRiskDistance = lowLevelRiskDistance
        self.resPath = self.resLE.text()

        self.close()
        

        
    
        

    
