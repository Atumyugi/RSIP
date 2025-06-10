import sys
import os
import os.path as osp

from ui.shp_process_dialog.excelTransLineDialog import Ui_excelTransLineDialog

from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtWidgets import QApplication, QDialog, QHBoxLayout, QVBoxLayout
from PyQt5.QtGui import QFont,QIcon

from qgis.core import QgsProject,QgsMapLayer,QgsMapLayerType,QgsVectorLayer,QgsWkbTypes,QgsLayerTreeGroup,QgsRasterLayer,QgsRectangle,QgsMapSettings,QgsCoordinateReferenceSystem
from qgis.gui import QgsMapCanvas,QgsProjectionSelectionDialog

from qfluentwidgets import (CardWidget, setTheme, Theme, IconWidget, BodyLabel,SubtitleLabel, CaptionLabel, PushButton,
                            TransparentToolButton, RoundMenu, Action,MessageBox)
from qfluentwidgets import FluentIcon as FIF

from yoyiUtils.qtTriggeredCommon import qtTriggeredCommonDialog
from yoyiUtils.yoyiFile import checkTifList
from yoyiUtils.yoyiTranslate import yoyiTrans
import yoyirs_rc

from openpyxl import Workbook,load_workbook

PROJECT = QgsProject.instance()

class ExcelTransLineDialogClass(Ui_excelTransLineDialog,QDialog):
    def __init__(self,yoyiTrs:yoyiTrans,parent=None):
        super().__init__(parent)
        self.parentWindow = parent
        self.parentMapCanvas : QgsMapCanvas = self.parentWindow.mapCanvas
        self.yoyiTrs = yoyiTrs
        self.setupUi(self)
        self.initMember()
        self.initUI()
        self.connectFunc()
    
    def initMember(self):
        self.pointExcelPath = None
        self.lineNameField = None
        self.pointIdField = None
        self.xField = None
        self.YField = None
        self.resPath = None
    
    def initUI(self):
        self.setWindowFlags(self.windowFlags() &~ Qt.WindowContextHelpButtonHint)

        self.selectSaveFilePB.setIcon(FIF.MORE)
        self.selectExcelFilePB.setIcon(FIF.MORE)

        self.refreshFieldView()

    def connectFunc(self):
        self.excelLE.textChanged.connect(self.refreshFieldView)

        self.selectExcelFilePB.clicked.connect(lambda: qtTriggeredCommonDialog().addFileTriggered(self.excelLE,'excel',parent=self))
        self.selectSaveFilePB.clicked.connect(lambda: qtTriggeredCommonDialog().selectSaveFileTriggered(self.resLE,'shp',parent=self))
        self.runPB.clicked.connect(self.runPBClicked)
        self.cancelPB.clicked.connect(self.close)

    def refreshFieldView(self):
        excelPath  = self.excelLE.text()
        print(excelPath)
        self.lineNameFieldCb.clear()
        if osp.exists(excelPath) and osp.splitext(excelPath)[1] == ".xlsx":
            
            pointWb = load_workbook(excelPath,read_only=True)
            pointWs = pointWb.active
            firstRow = next(pointWs.iter_rows(values_only=True,min_row=1, max_row=1))
            lineNameIndex = None
            pointIdIndex = None
            xIndex = None
            yIndex = None
            for index,rowContent in enumerate(firstRow):
                self.lineNameFieldCb.addItem(rowContent)
                self.pointIdFieldCb.addItem(rowContent)
                self.XFieldCb.addItem(rowContent)
                self.YFieldCb.addItem(rowContent)
                if rowContent == '线路名称':
                    lineNameIndex = index
                elif rowContent == '杆塔编号':
                    pointIdIndex = index
                elif rowContent == 'X' or rowContent == '经度' :
                    xIndex = index
                elif rowContent == 'Y' or rowContent == '纬度' :
                    yIndex = index
            
            if lineNameIndex:
                self.lineNameFieldCb.setCurrentIndex(lineNameIndex)
            if pointIdIndex:
                self.pointIdFieldCb.setCurrentIndex(pointIdIndex)
            if xIndex:
                self.XFieldCb.setCurrentIndex(xIndex)
            if yIndex:
                self.YFieldCb.setCurrentIndex(yIndex)


    def runPBClicked(self):
        if self.excelLE.text() == "" or self.resLE.text() == "":
            MessageBox(self.yoyiTrs._translate('警告'), self.yoyiTrs._translate('地址非法'), self).exec_()
            return
        
        if self.lineNameFieldCb.count() == 0:
            MessageBox(self.yoyiTrs._translate('警告'), 'excel行数为0', self).exec_()
            return
        
        self.pointExcelPath = self.excelLE.text()
        self.lineNameField = self.lineNameFieldCb.currentText()
        self.pointIdField = self.pointIdFieldCb.currentText()
        self.xField = self.XFieldCb.currentText()
        self.YField = self.YFieldCb.currentText()
        self.resPath = self.resLE.text()

        self.close()
        

        
    
        

    
